from django.core.validators import FileExtensionValidator
from django.db import models
import io
#try:
from PIL import Image
#except:
    #import Image
from pdf2image import convert_from_path
import pytesseract
import datetime

class UploadImage(models.Model):
    image = models.ImageField(upload_to='img/')
    result = models.ImageField(upload_to='result/')

    def transform(self, angle, gray):

        # アップロードされたファイルから画像オブジェクト生成
        org_img = Image.open(self.image)

        # PILでの画像処理ここから！
        ret_img = org_img.rotate(angle)

        if gray:
            ret_img = ret_img.convert('L')
        # PILでの画像処理ここまで！

        # 画像処理後の画像のデータをbufferに保存
        buffer = io.BytesIO()
        ret_img.save(fp=buffer, format=org_img.format)

        # 以前保存した画像処理後の画像ファイルを削除
        self.result.delete()

        # bufferのデータをファイルとして保存（レコードの更新も行われる）
        self.result.save(name=self.image.name, content=buffer)

class UploadPdf(models.Model):
    description = models.CharField(max_length=255, blank=True)
    pdfdocument = models.FileField(upload_to='pdf/')
    uploaded_at = models.DateTimeField(auto_now_add=True)


#
#https://stackoverflow.com/questions/66069902/in-django-how-to-convert-an-uploaded-pdf-file-to-an-image-file-and-save-to-the-c
#
from django.core.validators import FileExtensionValidator
from django.db.models.signals import post_save
from pdf2image import convert_from_path
from django.conf import settings
import os
from django.http import HttpResponse, JsonResponse  # Add

COVER_PAGE_DIRECTORY = 'coverdirectory/'
PDF_DIRECTORY = 'pdfdirectory/'
COVER_PAGE_FORMAT = 'jpg'

# this function is used to rename the pdf to the name specified by filename field
def set_pdf_file_name(instance, filename):
    return os.path.join(PDF_DIRECTORY, '{}.pdf'.format(instance.filename))

# not used in this example
def set_cover_file_name(instance, filename):
    return os.path.join(COVER_PAGE_DIRECTORY, '{}.{}'.format(instance.filename, COVER_PAGE_FORMAT))

class Pdffile(models.Model):
    # validator checks file is pdf when form submitted
    pdf = models.FileField(
        upload_to=set_pdf_file_name, 
        validators=[FileExtensionValidator(allowed_extensions=['pdf'])]
        )
    filename = models.CharField(max_length=20)            # save as file name
    pagenumforcover = models.IntegerField()
    coverpage = models.FileField(upload_to=set_cover_file_name)     # converted image file name
    result = models.ImageField(upload_to='ocrresult/')
    text = models.TextField(null=True)
    #xlsfile = models.FileField(upload_to='excel/')

    def pdftransform(self, angle, gray):

        # アップロードされたファイルから画像オブジェクト生成
        org_img = Image.open(self.coverpage)
        #result_img = Image.open(self.result)

        # PILでの画像処理ここから！
        ret_img = org_img.rotate(angle, expand=True)

        if gray:
            ret_img = ret_img.convert('L')
        # PILでの画像処理ここまで！
        
        # tesseract ocr   #ADD 3 lines
        import pyocr
        ocr_data={}
        #'''
        tesseract_path = "C:\Program Files\Tesseract-OCR"
        if tesseract_path not in os.environ["PATH"].split(os.pathsep):
            os.environ["PATH"] += os.pathsep+tesseract_path
            
        pyocr.tesseract.TESSERACT_CMD = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        tools = pyocr.get_available_tools()

        builder = pyocr.builders.LineBoxBuilder(tesseract_layout=6)
        #document = tools[0].image_to_string(img, lang="jpn", builder=builder)
        document = tools[0].image_to_string(ret_img, lang="jpn")
        document.replace("|", "")
        document.replace(", ", ",")
        document.replace("],", "1,")
        ocr_data['document'] = document
        self.text = document
        #'''
        # インストール方法：py -m pip install OpenPyXL
        import openpyxl as px
        #from openpyxl.drawing.image import Image
        # 新規ファイル作成
        #wb = px.load_workbook(r'media/excel/sample.xlsx')
        wb=px.Workbook()
        # 対象のシートを選択する
        ws = wb.active
        #ws = wb["Sheet"]
        # セル操作
        ## A1形式
        ws['A1'].value = '文字認識結果'
        ## R1C1形式
        #ws.cell(row=2, column=3).value = filename
        row=4
        column=2
        
        document=document.split("\n")
        i=1
        nrow=len(document)
        for text in document:
            text.replace('l', '')
            text.replace(', ', ',')
            text.replace('],','1,')
            ws.cell(row=row, column=1).value = i
            text=text.split()
            j=2
            for t in text:
                ws.cell(row=row, column=j).value=t
                j=j+1
            i=i+1
            row=row+1
        # 保存(絶対パス)
        dt=datetime.datetime.now()
        dtime=dt.strftime('%Y%m%d-%H%M%S')
        xlsfile='media/excel/result' + dtime+ '.xlsx'
        wb.save(xlsfile)
        
        '''
        #ocr_content = pytesseract.image_to_string(ret_img, lang='jpn')
        #ocr_data['content'] = ocr_content
        response_data = {}
        document= pytesseract.image_to_string(ret_img, lang="jpn")
        #document=document.split('|')
        ocr_data['document'] = document
        self.text = document
        '''        

        # 画像処理後の画像のデータをbufferに保存
        buffer = io.BytesIO()
        ret_img.save(fp=buffer, format=org_img.format)

        # 以前保存した画像処理後の画像ファイルを削除
        self.result.delete()

        # bufferのデータをファイルとして保存（レコードの更新も行われる）
        self.result.save(name=self.coverpage.name, content=buffer)

        return xlsfile

        #return HttpResponse(ocr_data, 'ocrz_list.html')
    

def convert_pdf_to_image(sender, instance, created, **kwargs):
    if created:
        # check if COVER_PAGE_DIRECTORY exists, create it if it doesn't
        # have to do this because of setting coverpage attribute of instance programmatically
        cover_page_dir = os.path.join(settings.MEDIA_ROOT, COVER_PAGE_DIRECTORY)

        if not os.path.exists(cover_page_dir):
            os.mkdir(cover_page_dir)

        # convert page cover (in this case) to jpg and save
        cover_page_image = convert_from_path(
            pdf_path=instance.pdf.path,
            dpi=200, 
            first_page=instance.pagenumforcover, 
            last_page=instance.pagenumforcover, 
            fmt=COVER_PAGE_FORMAT, 
            output_folder=cover_page_dir,
            )[0]
        cover_page_image.close()

        # get name of pdf_file 
        pdf_filename, extension = os.path.splitext(os.path.basename(instance.pdf.name))
        new_cover_page_path = '{}.{}'.format(os.path.join(cover_page_dir, pdf_filename), COVER_PAGE_FORMAT)
        
        # rename the file that was saved to be the same as the pdf file
        os.rename(cover_page_image.filename, new_cover_page_path)
        
        # get the relative path to the cover page to store in model
        new_cover_page_path_relative = '{}.{}'.format(os.path.join(COVER_PAGE_DIRECTORY, pdf_filename), COVER_PAGE_FORMAT)
        instance.coverpage = new_cover_page_path_relative

        # call save on the model instance to update database record
        instance.save()

post_save.connect(convert_pdf_to_image, sender=Pdffile)
